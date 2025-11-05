import sys
import os
import datetime
import holidays
import numpy as np
import pandas as pd
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.data_utils import fetch_latest_data, merge_yf_and_local
from utils.features import add_features
from utils.forecast_utils import (
    load_model,
    forecast_future_incremental,
    evaluate_model_performance,
    get_model_feature_names,
)

# ===============================
# CEK LIBUR NASIONAL & AKHIR PEKAN
# ===============================
today = datetime.date.today()
id_holidays = holidays.Indonesia()

if today.weekday() >= 5 or today in id_holidays:
    reason = "akhir pekan" if today.weekday() >= 5 else "libur nasional"
    print(f"⏸️ Hari ini ({today}) adalah {reason}. Prediksi otomatis dilewati.")
    sys.exit(0)

# ===============================
# KONFIGURASI
# ===============================
LOCAL_DATASET1_PATH = "data/IHSG 1993-2013.csv"
LOCAL_DATASET2_PATH = "data/IHSG 2013-2025⁄9⁄26.csv"
MODEL_PATH = "model/xgboost_ihsg_model.pkl"
OUTPUT_PATH = "./ihsg_forecast.xlsx"
EVAL_PATH = "./model_evaluation.xlsx"
N_FORECAST = 30

# ===============================
# MAIN FLOW
# ===============================
if __name__ == "__main__":
    print(f"🚀 Menjalankan update_forecast.py pada {today}")

    # 1. Ambil data terbaru
    print("📥 Mengambil data IHSG terbaru dari Yahoo Finance...")
    df_yf = fetch_latest_data()

    # 2. Baca dataset historis lokal
    print("📂 Membaca dataset historis lokal...")
    df1 = pd.read_csv(LOCAL_DATASET1_PATH)
    df2 = pd.read_csv(LOCAL_DATASET2_PATH)

    # 3. Gabungkan semuanya
    print("🧩 Menggabungkan semua dataset...")
    df_all_raw = merge_yf_and_local(df_yf, df1, df2)

    # 4. Feature engineering
    print("⚙️ Menambahkan fitur teknikal...")
    df_all = add_features(df_all_raw)

    # 5. Load model
    print("🧠 Memuat model XGBoost...")
    model = load_model(MODEL_PATH)

    # 6. Baca file Excel yang sudah ada (kalau ada)
    existing_df = None
    if os.path.exists(OUTPUT_PATH):
        print(f"📖 Membaca file existing: {OUTPUT_PATH}")
        try:
            existing_df = pd.read_excel(OUTPUT_PATH, sheet_name="Sheet1")
            # <-- NORMALISASI: convert to datetime AND drop any time-of-day by normalizing to midnight
            existing_df['Tanggal'] = pd.to_datetime(
                existing_df['Tanggal'], dayfirst=True, errors='coerce'
            ).dt.normalize()
            print(f"   ✓ File existing berisi {len(existing_df)} baris")
        except Exception as e:
            print(f"   ⚠️ Error membaca file existing: {e}")
            existing_df = None

    # 7. Generate prediksi incremental
    print("🔮 Membuat prediksi incremental...")
    updated_df = forecast_future_incremental(model, df_all, df_all_raw, existing_df, N_FORECAST)

    # 8. EVALUASI MENGGUNAKAN PREDIKSI HISTORIS
    print("\n📊 Mengevaluasi performa model dengan prediksi historis...")
    
    if existing_df is not None and len(existing_df) > 0:
        # Ambil periode 30 hari sejak kemarin (nilai aktual terakhir)
        today_dt = pd.Timestamp(today).normalize()   # pastikan midnight
        yesterday = today_dt - pd.Timedelta(days=1)
        date_30_days_ago = yesterday - pd.Timedelta(days=30)

        # Filter prediksi historis untuk 30 hari terakhir yang punya nilai aktual
        # Ambil data sampai kemarin saja (bandingkan tanpa komponen waktu)
        historical_data = (
            existing_df[existing_df['Tanggal'].dt.normalize() <= yesterday]
            .sort_values("Tanggal")
            .tail(90)                      # ambil buffer 90 hari
            .dropna(subset=["Terakhir"])  # pastikan hanya baris yang punya actual
            .tail(30)                      # ambil 30 valid
            .copy()
        )
        
        if len(historical_data) > 0:
            print(f"   📅 Periode evaluasi: {date_30_days_ago.date()} - {today_dt.date()}")
            print(f"   📈 Ditemukan {len(historical_data)} data historis dengan nilai aktual")
            
            # Debug: cek kolom yang ada
            print(f"   📋 Kolom yang tersedia: {historical_data.columns.tolist()}")
            
            # Bersihkan format angka pada kolom Terakhir (Prediksi)
            # Contoh: "8.164,80" -> 8164.80
            if 'Terakhir (Prediksi)' in historical_data.columns:
                def parse_number(x):
                    if pd.isna(x) or x == '' or x == ' ':
                        return np.nan
                    # Jika sudah float/int, langsung return
                    if isinstance(x, (int, float)):
                        return float(x)
                    # Jika string, bersihkan
                    x_str = str(x).strip()
                    # Hapus titik ribuan, ganti koma dengan titik
                    x_str = x_str.replace('.', '').replace(',', '.')
                    try:
                        return float(x_str)
                    except:
                        return np.nan
                
                historical_data['Predicted_Price'] = historical_data['Terakhir (Prediksi)'].apply(parse_number)
                
                # Debug: tampilkan beberapa contoh konversi
                print("\n   🔍 Contoh konversi Predicted_Price:")
                sample = historical_data[['Tanggal', 'Terakhir (Prediksi)', 'Predicted_Price']].head()
                for _, row in sample.iterrows():
                    print(f"      {row['Tanggal'].date()}: '{row['Terakhir (Prediksi)']}' -> {row['Predicted_Price']}")
            else:
                print("   ⚠️ Kolom 'Terakhir (Prediksi)' tidak ditemukan")
                historical_data['Predicted_Price'] = np.nan
            
            # Hitung Predicted_Return
            # Predicted_Return = (Predicted_Price / Terakhir_sebelumnya) - 1
            historical_data['Terakhir_prev'] = historical_data['Terakhir'].shift(1)
            historical_data['Predicted_Return'] = (
                historical_data['Predicted_Price'] / historical_data['Terakhir_prev']
            ) - 1
            
            # Hitung Actual Return
            historical_data['Actual_Return'] = (
                historical_data['Terakhir'] / historical_data['Terakhir_prev']
            ) - 1
            
            # Debug: tampilkan beberapa contoh perhitungan
            print("\n   🔍 Contoh perhitungan Return:")
            sample = historical_data[['Tanggal', 'Terakhir_prev', 'Terakhir', 'Predicted_Price', 'Actual_Return', 'Predicted_Return']].head(5)
            for _, row in sample.iterrows():
                print(f"      {row['Tanggal'].date()}:")
                print(f"         Terakhir prev: {row['Terakhir_prev']:.2f}")
                print(f"         Terakhir: {row['Terakhir']:.2f}")
                print(f"         Predicted: {row['Predicted_Price']:.2f}")
                print(f"         Actual Return: {row['Actual_Return']:.6f}")
                print(f"         Predicted Return: {row['Predicted_Return']:.6f}")
            
            # Filter baris yang valid (tidak NaN)
            valid_data = historical_data[
                historical_data['Predicted_Return'].notna() & 
                historical_data['Actual_Return'].notna() &
                historical_data['Predicted_Price'].notna()
            ].copy()
            
            if len(valid_data) > 0:
                print(f"   ✓ Berhasil memproses {len(historical_data)} data untuk evaluasi")
                
                # Evaluasi Return
                y_actual_return = valid_data['Actual_Return'].values
                y_pred_return = valid_data['Predicted_Return'].values
                
                return_mse = mean_squared_error(y_actual_return, y_pred_return)
                return_rmse = np.sqrt(return_mse)
                return_mae = mean_absolute_error(y_actual_return, y_pred_return)
                return_r2 = r2_score(y_actual_return, y_pred_return)
                
                # MAPE Return
                mask = y_actual_return != 0
                return_mape = np.mean(np.abs((y_actual_return[mask] - y_pred_return[mask]) / y_actual_return[mask]))
                
                print("\n📈 Evaluasi Return:")
                print(f"   MSE  : {return_mse:.8f}")
                print(f"   RMSE : {return_rmse:.6f}")
                print(f"   MAE  : {return_mae:.6f}")
                print(f"   R²   : {return_r2:.4f}")
                print(f"   MAPE : {return_mape:.2%}")
                
                # Evaluasi Harga
                y_actual_price = valid_data['Terakhir'].values
                y_pred_price = valid_data['Predicted_Price'].values
                
                price_mse = mean_squared_error(y_actual_price, y_pred_price)
                price_rmse = np.sqrt(price_mse)
                price_mae = mean_absolute_error(y_actual_price, y_pred_price)
                price_r2 = r2_score(y_actual_price, y_pred_price)
                
                # MAPE Harga
                price_mape = np.mean(np.abs((y_actual_price - y_pred_price) / y_actual_price))
                
                print("\n💰 Evaluasi Harga:")
                print(f"   MSE  : {price_mse:.2f}")
                print(f"   RMSE : {price_rmse:.2f}")
                print(f"   MAE  : {price_mae:.2f}")
                print(f"   R²   : {price_r2:.4f}")
                print(f"   MAPE : {price_mape:.2%}")
                
                # 9. Simpan hasil evaluasi (struktur asli)
                today_date = pd.Timestamp.today().date()
                
                eval_data = {
                    "Tanggal_Update": [today_date],
                    "RMSE": [return_rmse],
                    "MAPE": [return_mape],
                    "R2": [return_r2],
                    "MSE": [return_mse],
                    "Jumlah_Data_Evaluasi": [len(historical_data)],
                    "MSE_Price": [price_mse],
                    "RMSE_Price": [price_rmse],
                    "MAE_Price": [price_mae],
                    "R2_Price": [price_r2],
                }
                
                df_eval = pd.DataFrame(eval_data)
                
                # Simpan ke Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Evaluasi"
                
                for r in dataframe_to_rows(df_eval, index=False, header=True):
                    ws.append(r)
                
                # Format angka
                for row in ws.iter_rows(min_row=2):
                    for i, cell in enumerate(row):
                        if i == 0:  # Tanggal_Update
                            if isinstance(cell.value, (datetime.date, pd.Timestamp)):
                                cell.number_format = 'DD/MM/YYYY'
                        elif isinstance(cell.value, (int, float)):
                            if i == 5:  # Jumlah_Data_Evaluasi
                                cell.number_format = '0'
                            else:
                                cell.number_format = '[$-421]#,##0.00'
                
                wb.save(EVAL_PATH)
                print(f"\n💾 Hasil evaluasi disimpan ke {EVAL_PATH}")
            else:
                print("   ⚠️ Tidak ada data yang valid untuk evaluasi setelah filtering")
        else:
            print("   ⚠️ Tidak ditemukan data historis dengan nilai aktual dalam 30 hari terakhir")
    else:
        print("   ⚠️ File existing tidak ditemukan, evaluasi dilewati")

    # 10. Simpan hasil forecast
    print("\n📂 Menyimpan hasil prediksi ke Excel...")
    
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "Sheet1"

    for r in dataframe_to_rows(updated_df, index=False, header=True):
        ws_output.append(r)

    # Format kolom
    for row in ws_output.iter_rows(min_row=2):
        # Tanggal
        row[0].number_format = 'DD/MM/YYYY'
        
        # Kolom angka (B-E dan H)
        for i in [1, 2, 3, 4, 7]:
            cell = row[i]
            try:
                if cell.value not in (None, '', ' '):
                    cell.value = float(cell.value)
                    cell.number_format = '[$-421]#,##0.00'
            except Exception:
                pass

    wb_output.save(OUTPUT_PATH)
    print(f"💾 File tersimpan: {OUTPUT_PATH}")

    print("\n🎯 Proses update selesai sepenuhnya.")