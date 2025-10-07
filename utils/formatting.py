import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from datetime import timedelta
import holidays
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def fmt_price_eu(v):
    if pd.isna(v):
        return ""
    v = round(v, 2)  # dua desimal
    s = f"{v:,.2f}"  # default: 1,234.56
    # swap ke format CSV lokal: '.' ribuan, ',' desimal
    s = s.replace(',', 'X').replace('.', ',').replace('X', '.')
    return s

def fmt_volume_eu(v):
    if pd.isna(v):
        return ""
    try:
        val = float(v)
    except:
        return ""
    if abs(val) >= 1e9:
        out = val / 1e9
        s = f"{out:.2f}".replace('.', ',') + 'B'
    elif abs(val) >= 1e6:
        out = val / 1e6
        s = f"{out:.2f}".replace('.', ',') + 'M'
    else:
        return fmt_price_eu(val)
    return s

def save_to_excel(forecast_df, model_input_df, df_all_raw, output_path, skip_holidays=True):
    """
    Save historis + prediksi ke satu Excel sheet (Sheet1) dan simpan model_input_df ke sheet terpisah.
    Kolom Terakhir, Pembukaan, Tertinggi, Terendah disimpan sebagai angka (Number) di Excel.
    """

    def fmt_volume(x):
        if pd.isna(x):
            return ""
        try:
            v = float(x)
        except Exception:
            return ""
        if abs(v) >= 1e9:
            return f"{v/1e9:.2f}B"
        elif abs(v) >= 1e6:
            return f"{v/1e6:.2f}M"
        else:
            return f"{v:.2f}"

    # --- Siapkan data historis ---
    df_hist = df_all_raw.copy()
    df_hist['Tanggal'] = pd.to_datetime(df_hist.get('Tanggal'), dayfirst=True, errors='coerce')
    df_hist = df_hist.dropna(subset=['Tanggal']).sort_values('Tanggal').reset_index(drop=True)
    if df_hist.empty:
        raise RuntimeError("Data historis kosong setelah parsing tanggal.")
    last_hist_date = df_hist['Tanggal'].max().normalize()

    # --- Siapkan data prediksi ---
    fc = forecast_df.copy().reset_index(drop=True)
    if 'Tanggal' in fc.columns:
        fc['Tanggal'] = pd.to_datetime(fc['Tanggal'], dayfirst=True, errors='coerce')

    need_regen = False
    if fc['Tanggal'].isna().all():
        need_regen = True
    else:
        if fc['Tanggal'].min() <= last_hist_date:
            need_regen = True

    if need_regen:
        n = len(fc)
        dates = []
        candidate = last_hist_date + timedelta(days=1)
        id_holidays = holidays.Indonesia() if skip_holidays else set()
        while len(dates) < n:
            if candidate.weekday() >= 5 or (skip_holidays and candidate in id_holidays):
                candidate += timedelta(days=1)
                continue
            dates.append(candidate)
            candidate += timedelta(days=1)
        fc['Tanggal'] = pd.to_datetime(dates)

    fc = fc[fc['Tanggal'] > last_hist_date].reset_index(drop=True)

    # --- Gabungkan historis + prediksi ---
    out_rows = []
    out_cols = ['Tanggal', 'Terakhir', 'Pembukaan', 'Tertinggi', 'Terendah', 'Vol.', 'Perubahan%', 'Status']

    for _, r in df_hist.iterrows():
        out_rows.append({
            'Tanggal': r['Tanggal'],
            'Terakhir': pd.to_numeric(r.get('Terakhir', np.nan), errors='coerce'),
            'Pembukaan': pd.to_numeric(r.get('Pembukaan', np.nan), errors='coerce'),
            'Tertinggi': pd.to_numeric(r.get('Tertinggi', np.nan), errors='coerce'),
            'Terendah': pd.to_numeric(r.get('Terendah', np.nan), errors='coerce'),
            'Vol.': fmt_volume(r.get('Vol.', np.nan)),
            'Perubahan%': (f"{r['Perubahan%']:.2f}%" if pd.notna(r.get('Perubahan%')) else ''),
            'Status': 'Historis'
        })

    for _, r in fc.iterrows():
        out_rows.append({
            'Tanggal': r['Tanggal'],
            'Terakhir': pd.to_numeric(r.get('Predicted_Price', np.nan), errors='coerce'),
            'Pembukaan': np.nan,
            'Tertinggi': np.nan,
            'Terendah': np.nan,
            'Vol.': '',
            'Perubahan%': '',
            'Status': 'Prediksi'
        })

    out_df = pd.DataFrame(out_rows, columns=out_cols)
    out_df = out_df.sort_values('Tanggal').reset_index(drop=True)

    # --- Simpan ke Excel ---
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"

    # tulis data
    for r in dataframe_to_rows(out_df, index=False, header=True):
        ws1.append(r)

    # set format kolom
    for row in ws1.iter_rows(min_row=2):
        # kolom tanggal
        row[0].number_format = 'DD/MM/YYYY'
        # kolom angka (pastikan tipe float & format Number)
        for i in [1, 2, 3, 4]:  # kolom B-E
            cell = row[i]
            try:
                if cell.value not in (None, '', ' '):
                    cell.value = float(cell.value)
                    cell.number_format = '[$-421]#,##0.00'  # akan terdeteksi sebagai Number di Excel
            except Exception:
                pass

    # sheet kedua
    ws2 = wb.create_sheet("model_input")
    try:
        mid = model_input_df.copy()
        if 'Tanggal' in mid.columns:
            mid['Tanggal'] = pd.to_datetime(mid['Tanggal'], errors='coerce').dt.strftime('%d/%m/%Y')
        for r in dataframe_to_rows(mid, index=False, header=True):
            ws2.append(r)
    except Exception:
        pass

    wb.save(output_path)
    print(f"💾 File tersimpan: {output_path} (Sheet1 + model_input)")